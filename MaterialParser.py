"""
A = Color ID
D = Color Name
P = Color value
R = Material type
"""

def getLegoMaterialInfo(xlsx_path):
    """Gets information from excel file based on concret column queries.

    Args:
        xlsx_path (string): File path

    Returns:
        dict: Python dict with the material information.
    """

    import openpyxl
    wb = openpyxl.load_workbook(filename = xlsx_path)

    main_sheet = wb["Table 1"]
    max_row = main_sheet.max_row
    max_column = main_sheet.max_row

    data = {}

    for id in range(1, max_column):
        color_id    = main_sheet["A" + str(id)].value
        color_name  = main_sheet["D" + str(id)].value
        color_value = main_sheet["P" + str(id)].value
        material    = main_sheet["R" + str(id)].value

        data[str(color_id)] = {"color_name":color_name, "color_value":color_value, "material":material}

    return data

my_legos_materials = getLegoMaterialInfo(r"G:\Unreal Projects\LegoScene\colour_chart.xlsx")
m_103 = my_legos_materials.get("103")
print m_103
print m_103.get("color_value")
print m_103.get("material")



